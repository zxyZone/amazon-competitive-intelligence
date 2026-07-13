#!/usr/bin/env python3
"""Normalize Amazon competitor exports into canonical JSON.

Supports CSV, TSV, JSON, and XLSX when openpyxl is available.

中文说明：
这个脚本把卖家精灵/SIF/Sorftime/手动表格里不同叫法的字段，统一映射成
Skill 内部使用的标准字段，比如 asin、price、reviews、total_units。
后续评分脚本只认标准字段，所以字段别名主要在这里维护。
"""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
from typing import Any


# 字段别名表：左边是 Skill 内部标准字段，右边是导入表里可能出现的列名。
# 后续如果发现卖家精灵/SIF/Sorftime 导出的中文列名不同，优先在这里补别名。
FIELD_ALIASES = {
    "asin": ["asin", "ASIN"],
    "title": ["title", "product title", "商品标题", "标题"],
    "brand": ["brand", "品牌"],
    "seller": ["seller", "seller name", "卖家", "店铺"],
    "price": ["price", "价格", "售价"],
    "coupon": ["coupon", "优惠券", "coupon amount"],
    "effective_price": ["effective price", "到手价", "成交价"],
    "rating": ["rating", "评分"],
    "reviews": ["reviews", "ratings", "评论数", "评分数"],
    "reviews_increasement": ["reviews_increasement", "新增评分数", "月新增评分数"],
    "total_units": ["total_units", "monthly units", "月销量", "销量"],
    "total_units_growth": ["total_units_growth", "月销量增长率", "销量增长率"],
    "total_amount": ["total_amount", "monthly revenue", "月销售额", "销售额"],
    "total_amount_growth": ["total_amount_growth", "月销售额增长率", "销售额增长率"],
    "bsr": ["bsr", "bsr rank", "大类排名", "bsr排名"],
    "bsr_rank_cv": ["bsr_rank_cv", "近7天BSR增长数"],
    "bsr_rank_cr": ["bsr_rank_cr", "近7天BSR增长率"],
    "available_date": ["available_date", "launch date", "上架时间"],
    "variations": ["variations", "变体数"],
    "keyword": ["keyword", "关键词"],
    "searches": ["searches", "月搜索量", "搜索量"],
    "purchases": ["purchases", "月购买量", "购买量"],
    "purchase_rate": ["purchase_rate", "购买率"],
    "rank_position": ["rank_position", "自然排名"],
    "ad_position": ["ad_position", "广告排名"],
    "bid": ["bid", "PPC竞价", "ppc bid"],
    "traffic_percentage": ["traffic_percentage", "流量占比"],
}


def normalize_header(value: str) -> str:
    """把表头转成宽松匹配格式，减少大小写、下划线、空格造成的匹配失败。"""
    return " ".join(str(value).strip().lower().replace("_", " ").split())


def alias_map(headers: list[str]) -> dict[str, str]:
    """根据输入表头找到可映射的标准字段。"""
    normalized = {normalize_header(h): h for h in headers}
    result = {}
    for canonical, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            key = normalize_header(alias)
            if key in normalized:
                result[canonical] = normalized[key]
                break
    return result


def clean_value(value: Any) -> Any:
    """清洗单元格值：去掉货币符号、逗号、百分号，并尽量转成数字。"""
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        if value == "":
            return None
        cleaned = value.replace("$", "").replace(",", "").replace("%", "")
        try:
            number = float(cleaned)
            if number.is_integer():
                return int(number)
            return number
        except ValueError:
            return value
    return value


def read_rows(path: Path) -> list[dict[str, Any]]:
    """读取 CSV/TSV/JSON/XLSX；XLSX 依赖 openpyxl，没有时提示先转 CSV。"""
    suffix = path.suffix.lower()
    if suffix == ".json":
        data = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(data, dict):
            data = data.get("rows") or data.get("data") or [data]
        return list(data)
    if suffix in {".csv", ".tsv"}:
        delimiter = "\t" if suffix == ".tsv" else ","
        with path.open(newline="", encoding="utf-8-sig") as handle:
            return list(csv.DictReader(handle, delimiter=delimiter))
    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise SystemExit("Install openpyxl or export the workbook as CSV first.") from exc
        workbook = load_workbook(path, data_only=True, read_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h or "").strip() for h in rows[0]]
        return [dict(zip(headers, row)) for row in rows[1:]]
    raise SystemExit(f"Unsupported file type: {path.suffix}")


def normalize_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """把原始行转换为标准字段行，同时保留 _raw 方便追溯原始证据。"""
    if not rows:
        return []
    headers = list(rows[0].keys())
    mapping = alias_map(headers)
    normalized_rows = []
    for row in rows:
        item = {}
        for canonical, source in mapping.items():
            item[canonical] = clean_value(row.get(source))
        item["_raw"] = row
        normalized_rows.append(item)
    return normalized_rows


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("input", type=Path)
    parser.add_argument("-o", "--output", type=Path, default=Path("normalized_competitors.json"))
    args = parser.parse_args()

    rows = read_rows(args.input)
    normalized = normalize_rows(rows)
    args.output.write_text(json.dumps(normalized, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {len(normalized)} normalized rows to {args.output}")


if __name__ == "__main__":
    main()
